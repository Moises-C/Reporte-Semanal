import os
import re
import fdb
import openpyxl
import threading
import tkinter.font as tkFont
import tkinter as tk
import tkinter.filedialog as tkF
from ConexionBD import *
from sys import maxsize
from turtle import width
from tkinter import ttk
from tkinter import Label
from tkinter import Entry
from tkinter import messagebox
from tkinter import PhotoImage
from tkinter import Button
from tkinter import Radiobutton
from tkinter import Checkbutton
from tkinter import Frame
from tktooltip import ToolTip
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side


class Reporte():

    FAS = (
        '1427',
        '1428',
        '1464',
        '1548',
        '1558',
        '1642',
        '1742',
        '1744',
        '1787',
        '2442')
    FIS = (
        '1318',
        '1320',
        '1543',
        '1557',
        '1614',
        '1856',
        '1926',
        '2445',
        '2446')
    FCM = ('9001', '1801', '1802', '1800', '2045')

    diccionarioDivisiones = {"FIS": FIS, "FAS": FAS, "FCM": FCM}
    listaDivisones = []

    parteQuery = ""

    def __init__(self, *args, **kwargs):

        self.ventana = tk.Tk()
        self.ventana.iconbitmap('src/icono.ico')
        self.ventana.title("Reporte Semanal")
        self.centrarVentana(308, 340)
        self.ventana.configure(bg="white")
        self.ventana.resizable(0, 0)
        self.click_btn = PhotoImage(file='src/reporte.png')
        self.crearComponentes()
        self.ventana.mainloop()

    def crearComponentes(self):

        self.divisionFASEstado = tk.BooleanVar()
        self.divisionFISEstado = tk.BooleanVar()
        self.divisionFCMEstado = tk.BooleanVar()
        self.servExt = tk.BooleanVar()
        self.tipoOperacion = tk.IntVar()

        self.creacionFrames()
        self.creacionEtiquetas()
        self.creacionOpciones()
        self.creacionIngresoFechas()
        self.creacionBotones()
        self.creacionTips()

    def crearRadioBoton(self, ventana, opcion, valor, almacenar, cx, cy):

        RBtn = Radiobutton(ventana,
                           text=opcion,
                           value=str(valor),
                           variable=almacenar)

        RBtn.place(x=cx, y=cy)

        return RBtn

    def crearBoton(self, ventana, imagen, comando, fondo, cX, cY):

        boton = Button(ventana,
                       image=imagen,
                       command=comando,
                       bg=fondo,
                       borderwidth=0)

        boton.place(x=cX, y=cY)

        return boton

    def crearFrame(self, ventana, hg, ancho, alto, fondo, cX, cY):

        frame = Frame(
            ventana,
            highlightbackground="gray",
            highlightthickness=hg,
            width=ancho,
            height=alto,
            bd=0,
            bg=fondo,
            relief="groove")

        frame.place(x=cX, y=cY)

        return frame

    def crearEtiqueta(self, contenedor, texto, fuente, fondo, cX, cY):

        etiqueta = Label(
            contenedor,
            text=texto,
            font=fuente,
            bg=fondo)

        etiqueta.place(
            x=cX,
            y=cY)

        return etiqueta

    def crearCalendario(self, ventana, modo, idioma, formato, cX, cY):

        fecha = DateEntry(
            ventana,
            selectmode=modo,
            locale=idioma,
            date_pattern=formato)

        fecha.place(x=cX, y=cY)

        return fecha

    def crearCheckButton(self, ventana, texto, almacenar, cX, cY):

        RBtn = Checkbutton(ventana,
                           text=texto,
                           variable=almacenar)

        RBtn.place(x=cX, y=cY)

        return RBtn

    def creacionFrames(self):

        self.fechas = self.crearFrame(
            self.ventana,
            0,
            250,
            60,
            "#F0EEEE",
            30, 95)

        self.divisionFrame = self.crearFrame(
            self.ventana,
            0,
            250,
            60,
            "#F0EEEE",
            30, 160)

        self.divisionTipoOperacion = self.crearFrame(
            self.ventana,
            0,
            250,
            70,
            "#F0EEEE",
            30, 20)

        self.servicios = self.crearFrame(
            self.ventana,
            0,
            250,
            50,
            "#F0EEEE",
            30, 225)

    def creacionEtiquetas(self):

        self.crearEtiqueta(
            self.fechas,
            "FECHA INICIAL",
            'Arial 9 bold',
            "#F0EEEE",
            15,
            5)

        self.crearEtiqueta(
            self.fechas,
            "FECHA FINAL",
            'Arial 9 bold',
            "#F0EEEE",
            155,
            5)

        self.crearEtiqueta(
            self.fechas,
            "al",
            'Arial 9 bold',
            "#F0EEEE",
            120,
            30)

        self.crearEtiqueta(
            self.divisionFrame,
            "DIVSIONES",
            'Arial 9 bold',
            "#F0EEEE",
            85, 5)

        self.crearEtiqueta(
            self.divisionTipoOperacion,
            "OPERACIÓN",
            'Arial 9 bold',
            "#F0EEEE",
            85,
            5)

        self.crearEtiqueta(
            self.servicios,
            "SERVICIOS",
            'Arial 9 bold',
            "#F0EEEE",
            85,
            0)

    def creacionOpciones(self):

        self.divisionFAS = self.crearCheckButton(self.divisionFrame,
                                                 "FAS",
                                                 self.divisionFASEstado,
                                                 30, 25)

        self.divisionFIS = self.crearCheckButton(self.divisionFrame,
                                                 "FIS",
                                                 self.divisionFISEstado,
                                                 100, 25)

        self.divisionFCM = self.crearCheckButton(self.divisionFrame,
                                                 "FCM",
                                                 self.divisionFCMEstado,
                                                 170, 25)

        self.crearRadioBoton(
            self.divisionTipoOperacion,
            "IMPORTACIÓN",
            "1",
            self.tipoOperacion, 15, 25)

        self.crearRadioBoton(
            self.divisionTipoOperacion,
            "EXPORTACIÓN",
            "2",
            self.tipoOperacion, 140, 25)

        self.crearRadioBoton(
            self.divisionTipoOperacion,
            "AMBAS",
            "-1",
            self.tipoOperacion, 95, 45)

        self.servicioExtra = self.crearCheckButton(
            self.servicios,
            "S.E.",
            self.servExt,
            100, 20)

    def creacionIngresoFechas(self):

        self.fechaPagoInicial = self.crearCalendario(
            self.fechas,
            'day',
            "es",
            "dd/mm/yyyy",
            10, 30)

        self.fechaPagoFinal = self.crearCalendario(
            self.fechas,
            'day',
            "es",
            "dd/mm/yyyy",
            150, 30)

    def creacionTips(self):

        self.controlTip(
            self.boton,
            "Generar reporte en Excel.")

        self.controlTip(
            self.divisionFrame,
            "(OPCIONAL) Si quieres todas las divisiones, selecciona todas o no selecciones ninguna opción.")

        self.controlTip(
            self.servicios,
            "(OPCIONAL) Escoge algún servicio extra.")

    def creacionBotones(self):

        self.boton = self.crearBoton(
            self.ventana,
            self.click_btn,
            self.generar,
            "white", 135, 290)

    def estadoDivisiones(self):
        # Se agregan las divisiones elegidas a una lista

        self.listaDivisones.clear()

        self.getListadoDivisiones('FAS', self.divisionFASEstado.get())
        self.getListadoDivisiones('FIS', self.divisionFISEstado.get())
        self.getListadoDivisiones('FCM', self.divisionFCMEstado.get())

        return self.getRestriccionDivision()

    def centrarVentana(self, ancho, alto):

        wtotal = self.ventana.winfo_screenwidth()
        htotal = self.ventana.winfo_screenheight()

        # Guardamos el largo y alto de la ventana (geometry)
        wventana = ancho
        hventana = alto

        # Aplicamos la siguiente formula para calcular donde debería
        # posicionarse
        pwidth = round(wtotal / 2 - wventana / 2)
        pheight = round(htotal / 2 - hventana / 2)

        #  Se lo aplicamos a la geometría de la ventana
        self.ventana.geometry(
            str(wventana) +
            "x" +
            str(hventana) +
            "+" +
            str(pwidth) +
            "+" +
            str(pheight))

    def getServiciosClausula(self):

        return " AND CTRAO_PEDIDO.NUM_TRAF LIKE '%S.E%'" if self.servExt.get() else ''

    def getServiciosUnion(self):

        return " INNER JOIN CTRAO_PEDIDO \
                 ON ( CTRAO_PEDIDO.NUM_REFE = SAAIO_PEDIME.NUM_REFE)" if self.servExt.get() else ''

    def getRestriccionDivision(self):

        if len(self.listaDivisones) == 0:

            return ""

        parteQuery = " AND ("

        for indice in range(0, len(self.listaDivisones)):

            if indice == 0:

                parteQuery = parteQuery + " CTRAO_EMBAR.DIV_EMPR  IN" + \
                    str(self.diccionarioDivisiones[self.listaDivisones[indice]])
            else:

                parteQuery = parteQuery + " OR CTRAO_EMBAR.DIV_EMPR  IN" + \
                    str(self.diccionarioDivisiones[self.listaDivisones[indice]])

        parteQuery = parteQuery + ") "

        return parteQuery

    def getListadoDivisiones(self, division, divisionEstado):

        if divisionEstado:

            self.listaDivisones.append(division)

    def controlTip(self, objeto, texto):

        ToolTip(objeto, msg=texto, delay=0.1)

    def queryReporte(self):

        query = """SELECT ctrac_client.NOM_IMP,
                   'AICM'                            AS PLAZA,
                   CASE
                     WHEN saaio_pedime.IMP_EXPO = '1' THEN 'IMPORTACION'
                     ELSE 'EXPORTACION'
                   END                               AS OPERACION,
                   saaio_pedime.FEC_PAGO,
                   saaio_pedime.ADU_DESP,
                   saaio_pedime.PAT_AGEN,
                   saaio_pedime.NUM_PEDI,
                   saaio_pedime.CVE_PEDI,
                   saaio_pedime.NUM_REFE,
                   saaio_pedime.TOT_PAGO             AS SUMA,
                   CASE
                     WHEN saaio_pedime.CVE_CNTA IN ( '1A', '8G' ) THEN 'CAR BBVA'
                     WHEN saaio_pedime.CVE_CNTA = '5E' THEN 'CAR SANTANDER'
                     WHEN saaio_pedime.CVE_CNTA = '2A' THEN 'CAR BANAMEX'
                     WHEN saaio_pedime.CVE_CNTA = 'BW' THEN 'PECE BANAMEX'
                     WHEN saaio_pedime.CVE_CNTA = '3G' THEN 'PECE HSBC'
                   END                               AS IMPUESTOS,
                   CASE
                     WHEN (SELECT ctrao_embar.DIV_EMPR FROM ctrao_embar WHERE SAAIO_PEDIME.NUM_REFE = ctrao_embar.NUM_REFE) IN ('1318', '1320', '1543', '1557', '1614', '1926', '2445', '2446') THEN 'FIS'
                     WHEN (SELECT ctrao_embar.DIV_EMPR FROM ctrao_embar WHERE SAAIO_PEDIME.NUM_REFE = ctrao_embar.NUM_REFE) IN ('1427', '1428', '1464', '1548', '1558', '1642', '1742', '1744', '1787', '2442') THEN 'FAS'
                     WHEN (SELECT ctrao_embar.DIV_EMPR FROM ctrao_embar WHERE SAAIO_PEDIME.NUM_REFE = ctrao_embar.NUM_REFE) IN ('9001', '1801', '1802', '1800', '2045') THEN 'FCM'
                     WHEN (
                             SELECT FIRST 1 NUM_PEDIDO
                             FROM   SAAIO_FACPAR
                             WHERE  SAAIO_PEDIME.NUM_REFE = NUM_REFE) IN ('1318', '1320', '1543', '1557', '1614', '1926', '2445', '2446') THEN 'FIS'
                    WHEN (
                             SELECT FIRST 1 NUM_PEDIDO
                             FROM   SAAIO_FACPAR
                             WHERE  SAAIO_PEDIME.NUM_REFE = NUM_REFE) IN ('1427', '1428', '1464', '1548', '1558', '1642', '1742', '1744', '1787', '2442') THEN 'FAS'
                    WHEN (
                             SELECT FIRST 1 NUM_PEDIDO
                             FROM   SAAIO_FACPAR
                             WHERE  SAAIO_PEDIME.NUM_REFE = NUM_REFE) IN ('9001', '1801', '1802', '1800', '2045') THEN 'FCM'
                   END                               AS DIVISION,
                   (SELECT CASE
                             WHEN Count(*) = 0 THEN ''
                             ELSE 'S.E'
                           END
                    FROM   ctrao_pedido
                    WHERE  NUM_REFE = saaio_pedime.NUM_REFE
                           AND NUM_TRAF LIKE '%S.E%')
                   || (SELECT CASE WHEN COUNT(GUIA)>0 THEN '/HAND CARRIER'
                              ELSE '' END
                       FROM   saaio_guias
                       WHERE  saaio_pedime.NUM_REFE = NUM_REFE
                              AND GUIA LIKE '%MHC%') AS OBSERVACIONES,
                   ''                                AS QR,
            CASE
                      WHEN (SELECT ctrao_embar.DIV_EMPR FROM ctrao_embar WHERE SAAIO_PEDIME.NUM_REFE = ctrao_embar.NUM_REFE) IS NULL or (SELECT ctrao_embar.DIV_EMPR FROM ctrao_embar WHERE SAAIO_PEDIME.NUM_REFE = ctrao_embar.NUM_REFE) = '' THEN
                                (
                                        SELECT FIRST 1 NUM_PEDIDO
                                        FROM   SAAIO_FACPAR
                                        WHERE  SAAIO_PEDIME.NUM_REFE = NUM_REFE)
                      ELSE (SELECT ctrao_embar.DIV_EMPR FROM ctrao_embar WHERE SAAIO_PEDIME.NUM_REFE = ctrao_embar.NUM_REFE)
            END AS PLANTA,
            CASE
                      WHEN ROUND(
                                   (
                                   SELECT FEC_ETAP
                                   FROM   CTRAO_ETAPAS
                                   WHERE  CVE_ETAP = '130'
                                   AND    SAAIO_PEDIME.NUM_REFE = NUM_REFE) -
                                 (
                                        SELECT FEC_ETAP
                                        FROM   CTRAO_ETAPAS
                                        WHERE  CVE_ETAP = '030'
                                        AND    SAAIO_PEDIME.NUM_REFE = NUM_REFE),0) >= 2 THEN 'ALERTA'
                                            || ROUND(
                                                       (
                                                       SELECT FEC_ETAP
                                                       FROM   CTRAO_ETAPAS
                                                       WHERE  CVE_ETAP = '130'
                                                       AND    SAAIO_PEDIME.NUM_REFE = NUM_REFE) -
                                                     (
                                                            SELECT FEC_ETAP
                                                            FROM   CTRAO_ETAPAS
                                                            WHERE  CVE_ETAP = '030'
                                                            AND    SAAIO_PEDIME.NUM_REFE = NUM_REFE),0)
                      ELSE ''
           END AS DIAS_ALMA
            FROM   saaio_pedime
                   INNER JOIN ctrac_client
                           ON ( ctrac_client.CVE_IMP = saaio_pedime.CVE_IMPO )
                   INNER JOIN CTRAO_EMBAR
                           ON (CTRAO_EMBAR.NUM_REFE = SAAIO_PEDIME.NUM_REFE)""" + \
            self.getServiciosUnion() + \
            """WHERE  saaio_pedime.CVE_IMPO = '1510'
                   AND saaio_pedime.ADU_DESP = '470'
                   AND saaio_pedime.PAT_AGEN = '1742' """  \
                   + self.validarOperacion() \
                   + self.getFechas() \
                   + self.estadoDivisiones() \
                   + self.getServiciosClausula() \
            + "ORDER BY SAAIO_PEDIME.FEC_PAGO "

        return query

    def queryRefesFaltantes(self):

        query = """SELECT
                        TMP.REFE_PAGADAS
                    FROM (SELECT NUM_REFE AS REFE_PAGADAS,
                         (SELECT NUM_REFE
                          FROM CTRAO_EMBAR
                          WHERE NUM_REFE = SAAIO_PEDIME.NUM_REFE) AS  REFE_QR
                          FROM   saaio_pedime
                          WHERE  CVE_IMPO = '1510'
                                 AND ADU_DESP = '470'
                                 AND PAT_AGEN = '1742' """ \
                                 + str(self.getFechas())  \
                                 + str(self.validarOperacion()) \
            + """ORDER BY FEC_PAGO)  TMP
                    WHERE TMP.REFE_QR IS NULL"""

        return query

    def getReferenciasSinRectificar(self):

        self.conexionBD = ConexionDB()
        self.conexionBD.establecerConexion('470')
        self.conexionBD.colocarEjecutarQuery(self.queryRefesFaltantes())

        refes = ""

        for referencias in self.conexionBD.getEjecucionQuery():

            refes += referencias[0] + ", "

        return re.sub(", $", "", refes)

    def validarOperacion(self):

        if self.tipoOperacion.get() == 0:

            messagebox.showwarning(
                "A D V E R T E N C I A",
                "Escoge el tipo de operación.")

            return ".-.-.-"

        elif self.tipoOperacion.get() > 0:

            return " AND SAAIO_PEDIME.IMP_EXPO = '" + \
                str(self.tipoOperacion.get()) + "'"

        else:

            return " "

    def generar(self):

        referencias = self.getReferenciasSinRectificar()
        operacion = self.tipoOperacion.get()

        if len(referencias) > 0:

            messagebox.showinfo(
                'Referencias no capturadas en sistema',
                "Referencias sin observaciones de rectificar: \n" +
                referencias)

        if operacion > -2 and operacion < 4:

            generarReporte = Archivo(self.queryReporte())

    def getFechas(self):

        fechaInicial = self.fechaPagoInicial.get_date().strftime("%Y-%m-%d") + " 00:00"
        fechaFinal = self.fechaPagoFinal.get_date().strftime("%Y-%m-%d") + " 23:59"

        return " AND SAAIO_PEDIME.DIA_PAGO BETWEEN '" + \
            fechaInicial + "' AND '" + fechaFinal + "' "

class Archivo:

    titulosImpo = ["CLIENTE", "PLAZA", "OPERACION",
                   "FECHA", "ADUANA", "PATENTE",
                   "PEDIMENTO", "CLAVE PEDIMENTO", "REFERENCIA", "SUMA",
                   "IMPUESTOS PAGADOS", "DIVISION",
                   "OBSERVACIONES", "QR", "PLANTA", "DIAS ALMA"]

    query = ""

    def __init__(self, query):

        self.con = fdb.connect(
            dsn='192.168.90.7:c:/CASAWIN/CARBD7/CASA.GDB',
            user='admin',
            password='admin')
        self.consulta = self.con.cursor()
        self.consultaPart = self.con.cursor()
        self.query = query
        self.my_wb = openpyxl.Workbook()
        self.my_sheet = self.my_wb.active
        self.colocarTitulos(self.titulosImpo)
        self.registros()
        self.guardar()

    def colocarTitulos(self, titulos):

        i = 0

        for titulo in titulos:

            self.colocarTitulo(1, i + 1, titulo)
            self.fondoCelda(1, i + 1)
            i += 1

    def registros(self):

        x = 2

        for fila in self.consulta.execute(self.query):

            for indice in range(0, len(fila) - 2):

                hilo1 = threading.Thread(
                    target=self.colocar, args=(
                        x, indice + 1, fila[indice],))
                hilo2 = threading.Thread(target=self.colocar, args=(
                    x, indice + 2, fila[indice + 1],))
                hilo3 = threading.Thread(target=self.colocar, args=(
                    x, indice + 3, fila[indice + 2],))

                if indice + 1 == 4:

                    self.colocarFormatoCelda(
                        x, indice + 1, fila[indice], "dd/mm/YYYY")

                hilo1.start()
                hilo2.start()
                hilo3.start()

            x += 1

    def obtenerValor(self, valorBuscado, cadena, valorBuscadoTmp):

        busqueda = re.search(valorBuscado, cadena)

        if busqueda:

            cadena = busqueda.group(0)
            cadena = re.sub("/+", "", cadena)
            cadena = re.sub(valorBuscadoTmp, "", cadena)

        else:

            cadena = ""

        return cadena

    def colocarTitulo(self, x, y, valor):

        celda = self.my_sheet.cell(row=x, column=y)
        celda.value = valor
        celda.font = Font(bold=True, name='Calibri', size=11, color="000000")
        celda.alignment = Alignment(
            wrap_text=True,
            horizontal='center',
            vertical='center')

    def colocar(self, x, y, valor):

        if y == 13:
            valor = re.sub("^\\s+/", "", valor)
        celda = self.my_sheet.cell(row=x, column=y)
        celda.value = valor
        celda.font = Font(name='Calibri', size=11)

    def colocarFormatoCelda(self, x, y, valor, formato):

        celda = self.my_sheet.cell(row=x, column=y)
        celda.value = valor
        celda.font = Font(name='Calibri', size=11)
        celda.number_format = formato

    def fondoCelda(self, x, y):

        celda = self.my_sheet.cell(row=x, column=y)
        celda.fill = PatternFill("solid", fgColor="00B0F0")

    def guardar(self):

        try:

            files = [('Archivo de Excel', '.xlsx')]
            self.f = tkF.asksaveasfile(filetypes=files, defaultextension=files)
            self.my_wb.save("" + self.f.name + "")
            self.abrir()

        except BaseException:

            pass

    def abrir(self):

        if messagebox.askquestion("Finalizado",
                                  ("¿Deseas abrir el reporte?")) == 'yes':

            os.system("start EXCEL.EXE " + '"' + self.f.name + '"')

if __name__ == '__main__':

    iniciar = Reporte()
